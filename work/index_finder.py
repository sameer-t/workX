# -*- coding: utf-8 -*-
"""
Created on Wed Sep 24 14:29:17 2014

@author: sameer
"""

import re, collections
from openpyxl import load_workbook

def convert(data):
    if isinstance(data, basestring):
        return data.encode('utf-8')
    elif isinstance(data, collections.Mapping):
        return dict(map(convert, data.iteritems()))
    elif isinstance(data, collections.Iterable):
        return type(data)(map(convert, data))
    else:
        return data

pitest = open('pilist.txt','r')
pdflist =eval(pitest.readlines()[0])
pitest.close()

ok = '(-_-)'
  
l= 1
pi9 = load_workbook("/home/sameer/2008_input.xlsx")
pi = pi9.active
npl = []
for rowno in range(2,3578):
    tempd = dict()
    tempd["lname"] = convert(pi.cell(row = rowno, column = 2).value)
    tempd["pages"] = pi.cell(row = rowno, column = 8).value.split(',')
    npl.append(tempd)

for k in range(len(npl)):
    tname = npl[k]['lname']
    reg = r',([^,]*' + tname + r'[^,]*),'
    try:
        page1 = int(npl[k]['pages'][0])
        tpage1 = str(pdflist[page1-1]).replace('\n',' ')
        exps1 = re.findall(reg,tpage1)
        try:
            page2 = int(npl[k]['pages'][1])
            tpage2 = str(pdflist[page2-1]).replace('\n',' ')
            exps2 = re.findall(reg,tpage2)
            try:
                pi.cell(row = k+2, column = 9).value = exps1[0]
            except:
                print ok
            try:
                pi.cell(row = k+2, column = 10).value = exps2[0]
            except:
                print ok
            try:
                pi.cell(row = k+2, column = 11).value = str(exps1[1:])
            except:
                print ok
            try:
                pi.cell(row = k+2, column = 12).value = str(exps2[1:])
            except:
                print ok
        except:
            try:
                pi.cell(row = k+2, column = 9).value = exps1[0]
            except:
                print ok
            try:
                pi.cell(row = k+2, column = 10).value = exps1[1]
            except:
                print ok
            try:
                pi.cell(row = k+2, column = 11).value = exps1[2]
            except:
                print ok
            try:
                pi.cell(row = k+2, column = 12).value = str(exps1[3:])
            except:
                print ok
        print tname + ': ' + exps1[0]
    except:
        print tname
        

pi9.save('2008_Output2.xlsx')
    
#t1 = pdflist[257]
#t1 = str(t1)
#    .{0,15}Aaberg.{0,15}
#re.findall(r',([^,]*Aaberg[^,]*),',tpage)